# This Python file uses the following encoding: utf-8

# if __name__ == "__main__":
#     pass

import base64
import hmac
import os
import openpyxl
import copy
from urllib.parse import quote


def data_save_to_new_excel(file_path):
    os.chdir(file_path)
    if os.path.exists('token.xlsx'):
        # 返回一个workbook数据类型的值
        workbook = openpyxl.load_workbook('token.xlsx')
        # 获取活动表
        sheet = workbook.get_sheet_by_name("待烧录Token")
        print(sheet)
        # 获取行数
        token_rows = sheet.max_row

    else:
        # 返回一个workbook数据类型的值
        wb = openpyxl.Workbook()
        # 获取活动表
        sheet = wb.active
        sheet.title = "待烧录Token"
        new_sheet = wb.create_sheet("已烧录Token")
        wb.save('token.xlsx')
        sheet_names = wb.get_sheet_names
        print(sheet_names)


if __name__ == '__main__':
    path = r'D:\\lijianlin\\项目\\软件\\API\\数据集\\两客一危\\0723\\'
    data_save_to_new_excel(path)
