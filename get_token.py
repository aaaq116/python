import base64
import hmac
import os
import openpyxl
import copy
from urllib.parse import quote


def read_excel_data(file_path, file_name):
    data = []
    # 路径
    os.chdir(file_path)
    workbook = openpyxl.load_workbook(file_name)
    # 获取活动表
    sheet = workbook.active
    # 获取行数
    rows = sheet.max_row
    # 获取指定行列数据
    for i in range(rows):
        if i >= 1:
            # 获取指定行列数据
            # 获取设备名
            name = sheet.cell(row=i+1, column=1).value
            # 获取产品ID
            product_id = sheet.cell(row=i+1, column=4).value
            # 获取秘钥
            secret_key = sheet.cell(row=i+1, column=5).value
            # 拼接
            res_string = product_id + '/devices/' + name
            token = get_token(res_string, secret_key)
            data.append(token)
    data_save_to_new_excel(file_path, data)


def data_save_to_new_excel(file_path, data):
    os.chdir(file_path)
    if os.path.exists('token.xlsx'):
        # 返回一个workbook数据类型的值
        workbook = openpyxl.load_workbook('token.xlsx')
        # 获取活动表
        sheet = workbook.get_sheet_by_name("已烧录Token")
        # 获取活动表
        new_sheet = workbook.get_sheet_by_name("待烧录Token")
        # 获取行数
        token_rows = sheet.max_row
        data_rows = len(data)
        # 烧写列表
        write_data = copy.deepcopy(data)
        # 判断数据是否存在
        for i in range(data_rows):
            for row in range(2, token_rows + 1):
                if data[i] == sheet.cell(row=row, column=1).value:
                    print("该数据已存在烧录Token", data[i])
                    write_data.remove(data[i])
        for j in range(len(write_data)):
            new_sheet['A{}'.format(j + 2)] = write_data[j]
        workbook.save('token.xlsx')
    else:
        # 返回一个workbook数据类型的值
        wb = openpyxl.Workbook()
        # 获取活动表
        sheet = wb.active
        sheet.title = "待烧录Token"
        for i in range(len(data)):
            sheet['A1'] = 'Token'
            sheet['A{}'.format(i + 2)] = data[i]
        old_sheet = wb.create_sheet("已烧录Token")
        old_sheet['A1'] = 'Token'
        wb.save('token.xlsx')


def get_token(name, access_key):

    version = '2018-10-32'

    # res = 'mqs/%s' % id  # 通过实例名称访问MQ
    res = 'products/%s' % name  # 通过产品ID访问产品API

    # 用户自定义token过期时间
    # et = str(int(time.time()) + 3600)
    et = "1710382043"

    # 签名方法，支持md5、sha1、sha256
    method = 'sha1'

    # 对access_key进行decode
    decode_key = base64.b64decode(access_key)

    # 计算sign
    org = et + '\n' + method + '\n' + res + '\n' + version
    sign_b = hmac.new(key=decode_key, msg=org.encode(), digestmod=method)
    sign = base64.b64encode(sign_b.digest()).decode()

    # value 部分进行url编码，method/res/version值较为简单无需编码
    sign = quote(sign, safe='')
    res = quote(res, safe='')

    # token参数拼接
    token = 'version=%s&res=%s&et=%s&method=%s&sign=%s' % (version, res, et, method, sign)

    return token


if __name__ == '__main__':
    path = "C:/Users/cgn12/Desktop"
    file = "test.xlsx"
    read_excel_data(path, file)
